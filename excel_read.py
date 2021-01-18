# 导入读取excel库


import openpyxl
import os


# 读sheet
def read_sheet(sheet):
    rows_data = list(sheet.rows)
    # 获取表单的表头信息
    titles = []
    for title in rows_data[0]:
        titles.append(title.value)
    print(titles)
    # 获取表单的类型信息
    types = []
    for title in rows_data[2]:
        types.append(title.value)
    print(types)
    # 获取表单的前后端信息
    clientOrServers = []
    for title in rows_data[3]:
        clientOrServers.append(title.value)
    print(clientOrServers)

    end_str = "\"" + sheet.title + "\"" + ":{"

    for row in rows_data:
        for title in row:
            if title.row > 4 and clientOrServers[title.col_idx - 1] == 'c':
                if title.col_idx == 1:
                    end_str += "\"" + str(title.value) + "\"" + ":{"
                else:
                    if title.col_idx > 2:
                        end_str += ','
                    if types[title.col_idx - 1] == 'int':
                        end_str += "\"" + titles[title.col_idx - 1] + "\":" + str(title.value)
                    elif types[title.col_idx - 1] == 'str':
                        end_str += "\"" + titles[title.col_idx - 1] + "\":\"" + str(title.value) + "\""
                if title.col_idx == len(row):
                    end_str += "},"
    if end_str.endswith(','):
        end_str = end_str[:-1]
    end_str += "}"
    print(end_str)
    return end_str


# 读excel
def read_excel(path, excelName):
    excelPath = path + excelName;
    # 打开文件
    wb = openpyxl.load_workbook(excelPath)
    # 获取所有sheet的名字
    print(wb.sheetnames)

    names = excelName.split('.')
    js_str = '\"' + names[0] + '\":{'
    for sheet in wb.sheetnames:
        sheet_name = wb[sheet]
        js_str += read_sheet(sheet_name)
        js_str += ','
    if js_str.endswith(','):
        js_str = js_str[:-1]
    js_str += '}'
    print(js_str)
    return js_str


# 将所有excel的数据合到一个文件里
def readAllExcel():
    path = os.path.abspath('.')
    pathExcel = path + '/excel/'
    dirs = os.listdir(pathExcel)
    json = '{'
    for excel in dirs:
        if excel.endswith('xlsx') and excel.find('~') == -1:
            print('excel=', excel)
            json += read_excel(pathExcel, excel)
            json += ','
    if json.endswith(','):
        json = json[:-1]
    json += '}'
    print('json=', json)
    return json


# 更新json文件
def reWriteJson(json):
    filename = os.path.abspath('.') + '/test.json'
    print(filename)
    fd = open(filename, 'w')
    fd.write(json)
    fd.close()


jsonStr = readAllExcel()
reWriteJson(jsonStr)
